"""Servicios y utilidades para interactuar con terminales ZKTeco."""
from __future__ import annotations

import csv
import json
import logging
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Set, Tuple
from urllib.error import URLError
from urllib.parse import urlencode
from urllib.request import urlopen

import pymysql
from pymysql.cursors import DictCursor
from flask import make_response, send_file
from openpyxl import Workbook, load_workbook

from zk import const
from zk_tools import DEFAULT_PORT, connect_with_retries
from .config import get_setting

logger = logging.getLogger(__name__)

TERMINAL_EMPLOYEES: Dict[str, List[dict]] = {}
SELECTED_EMPLOYEES: Dict[str, Set[str]] = {}
TERMINAL_LIST_PATH = Path(__file__).resolve().parent.parent / "terminales.txt"

DATABASE_TERMINAL_KEY = "__database__"
DATABASE_TERMINAL_LABEL = "Empleados contratados RRHH"
ZKTIME_TERMINAL_KEY = "__zktime__"
ZKTIME_TERMINAL_LABEL = "Empleados ZK Time"

SPECIAL_TERMINALS: Dict[str, str] = {
    DATABASE_TERMINAL_KEY: DATABASE_TERMINAL_LABEL,
    ZKTIME_TERMINAL_KEY: ZKTIME_TERMINAL_LABEL,
}


def get_special_terminal_options() -> List[Dict[str, str]]:
    """Devuelve la lista de orígenes especiales disponibles."""
    return [{"value": key, "label": label} for key, label in SPECIAL_TERMINALS.items()]


def get_special_terminal_label(value: Optional[str]) -> Optional[str]:
    """Obtiene la etiqueta asociada a un origen especial."""
    if not value:
        return None
    return SPECIAL_TERMINALS.get(value)


def normalize_special_terminal_value(value: Optional[str]) -> Optional[str]:
    """Normaliza el valor recibido y comprueba si corresponde a un origen especial."""
    if not value:
        return None
    stripped = value.strip()
    return stripped if stripped in SPECIAL_TERMINALS else None


def is_special_terminal(value: Optional[str]) -> bool:
    """Indica si el valor corresponde a un origen especial."""
    return normalize_special_terminal_value(value) is not None


def _normalize_setting(value: Optional[str]) -> Optional[str]:
    """Devuelve el valor sin espacios o None si está vacío."""
    if value is None:
        return None
    stripped = str(value).strip()
    return stripped or None


def _normalize_numeric_identifier(value: Any) -> Optional[int]:
    """Convierte un identificador numérico en int eliminando ceros a la izquierda."""
    text = _normalize_setting(value)
    if not text or not text.isdigit():
        return None
    return int(text)


def _normalize_positive_numeric_text(value: Any) -> Optional[str]:
    """Devuelve una cadena numérica positiva (>0) sin ceros iniciales."""
    identifier = _normalize_numeric_identifier(value)
    if identifier is None:
        return None
    if identifier <= 0:
        return None
    return str(identifier)


RRHH_CARD_UPDATE_URL = _normalize_setting(get_setting("RRHH_CARD_UPDATE_URL"))
RRHH_CARD_UPDATE_TIMEOUT = 10


def _get_setting_any(keys: Sequence[str], default: Optional[str] = None) -> Optional[str]:
    """Obtiene el primer valor disponible entre varias claves posible."""
    for key in keys:
        value = _normalize_setting(get_setting(key))
        if value is not None:
            return value
    return default


def _format_table_reference(identifier: str) -> str:
    """Construye una referencia SQL segura para un identificador opcionalmente cualificado."""
    parts = [part.strip() for part in (identifier or "").split(".") if part.strip()]
    if not parts:
        raise ValueError("El nombre de la tabla de ZK Time no puede estar vacío.")

    sanitized_parts = []
    for part in parts:
        if not part.replace("_", "").isalnum():
            raise ValueError(
                f"El identificador '{identifier}' contiene caracteres no permitidos."
            )
        sanitized_parts.append(f"`{part}`")
    return ".".join(sanitized_parts)


def _normalize_zktime_employee_record(raw: Dict[str, Any]) -> Optional[dict]:
    """Convierte un registro de ZK Time en el formato usado por la aplicación."""
    if not raw:
        return None

    def pick(*candidates: str) -> str:
        for candidate in candidates:
            value = _normalize_setting(raw.get(candidate))
            if value:
                return value
        return ""

    code = pick("codigo", "CODIGO", "user_id", "USER_ID")
    if not code:
        return None

    surnames = pick("apellidos", "APELLIDOS")
    given_name = pick("nombre", "NOMBRE")
    alias = pick("alias", "ALIAS")
    nif = pick("nif", "NIF", "dni", "DNI")
    tarjeta = pick("tarjeta", "TARJETA", "card", "CARD")

    full_name_parts = [part for part in [surnames, given_name] if part]
    full_name = " ".join(full_name_parts).strip() or alias or code
    alias_value = alias or None
    nif_value = nif or None
    surnames_value = surnames or None
    given_name_value = given_name or None
    tarjeta_value = tarjeta or None

    return {
        "uid": code,
        "user_id": code,
        "name": full_name,
        "card": tarjeta_value or alias_value,
        "alias": alias_value,
        "nif": nif_value,
        "dni": nif_value,
        "surnames": surnames_value,
        "given_name": given_name_value,
        "tarjeta": tarjeta_value,
        "code": code,
        "privilege": "",
        "group_id": "",
        "biometrics": [],
        "last_seen": None,
        "contract_from": None,
        "medical_leave_from": None,
        "vacation_status": None,
    }


def _get_zktime_connection_params() -> Dict[str, Any]:
    """Obtiene la configuración necesaria para conectarse a ZK Time."""
    host = _get_setting_any(["zktime_host", "ZKTIME_HOST"])
    user = _get_setting_any(["zktime_user", "ZKTIME_USER"])
    password = _normalize_setting(get_setting("zktime_password")) or _normalize_setting(
        get_setting("ZKTIME_PASSWORD")
    )
    database = _get_setting_any(["zktime_database", "ZKTIME_DATABASE"])
    port_value = _get_setting_any(["zktime_port", "ZKTIME_PORT"])
    charset = _get_setting_any(["zktime_charset", "ZKTIME_CHARSET"], "utf8mb4")
    table = _get_setting_any(["zktime_table", "ZKTIME_TABLE"], "empleados001")
    query = _normalize_setting(
        get_setting("zktime_query")
    ) or _normalize_setting(get_setting("ZKTIME_QUERY"))

    missing = [
        name
        for name, value in {
            "zktime_host": host,
            "zktime_user": user,
            "zktime_password": password,
            "zktime_database": database,
        }.items()
        if not value
    ]
    if missing:
        raise RuntimeError(
            "Faltan parámetros para conectar con ZK Time: "
            + ", ".join(sorted(missing))
        )

    if port_value is None:
        port = 3306
    else:
        try:
            port = int(port_value)
        except (TypeError, ValueError) as exc:
            raise ValueError("El puerto zktime_port debe ser un número entero.") from exc

    formatted_table = _format_table_reference(table)

    return {
        "host": host,
        "user": user,
        "password": password,
        "database": database,
        "port": port,
        "charset": charset,
        "table": formatted_table,
        "query": query,
    }


def load_zktime_employees() -> List[dict]:
    """Consulta la tabla de empleados de ZK Time y devuelve registros normalizados."""
    params = _get_zktime_connection_params()
    query = params["query"] or (
        f"SELECT nif, apellidos, nombre, alias, codigo, tarjeta FROM {params['table']}"
    )

    connection = pymysql.connect(
        host=params["host"],
        user=params["user"],
        password=params["password"],
        database=params["database"],
        port=params["port"],
        charset=params["charset"],
        cursorclass=DictCursor,
        autocommit=True,
    )
    try:
        with connection.cursor() as cursor:
            cursor.execute(query)
            rows: Sequence[Dict[str, Any]] = cursor.fetchall()
    except pymysql.MySQLError as exc:
        logger.exception("Error al consultar ZK Time: %s", exc)
        raise RuntimeError(f"No se pudo consultar la base de datos ZK Time: {exc}") from exc
    finally:
        connection.close()

    normalized: List[dict] = []
    for row in rows:
        normalized_row = _normalize_zktime_employee_record(row)
        if normalized_row:
            normalized.append(normalized_row)
    return normalized


def refresh_zktime_cache() -> List[dict]:
    """Actualiza la caché interna con los empleados obtenidos desde ZK Time."""
    employees = load_zktime_employees()
    TERMINAL_EMPLOYEES[ZKTIME_TERMINAL_KEY] = employees
    SELECTED_EMPLOYEES.pop(ZKTIME_TERMINAL_KEY, None)
    return employees


def update_zktime_cards(employees: Iterable[dict]) -> Tuple[int, int, List[Tuple[int, str]]]:
    """Actualiza los números de tarjeta de los empleados proporcionados en ZK Time."""
    params = _get_zktime_connection_params()
    updates: List[Tuple[str, int]] = []
    updates_info: List[Tuple[int, str]] = []

    for employee in employees:
        numeric_user_id = _normalize_numeric_identifier(employee.get("user_id"))
        card_text = _normalize_positive_numeric_text(employee.get("tarjeta") or employee.get("card"))
        if numeric_user_id is None or card_text is None:
            continue
        updates.append((card_text, numeric_user_id))
        updates_info.append((numeric_user_id, card_text))

    if not updates:
        return 0, 0, []

    connection = pymysql.connect(
        host=params["host"],
        user=params["user"],
        password=params["password"],
        database=params["database"],
        port=params["port"],
        charset=params["charset"],
        cursorclass=DictCursor,
        autocommit=True,
    )
    try:
        with connection.cursor() as cursor:
            sql = f"UPDATE {params['table']} SET tarjeta = %s WHERE CAST(codigo AS UNSIGNED) = %s"
            affected = cursor.executemany(sql, updates)
    except pymysql.MySQLError as exc:
        logger.exception("Error al actualizar tarjetas en ZK Time: %s", exc)
        raise RuntimeError(f"No se pudieron actualizar las tarjetas en ZK Time: {exc}") from exc
    finally:
        connection.close()

    return affected or 0, len(updates), updates_info


def update_rrhh_cards(
    employees: Iterable[dict],
) -> Tuple[int, int, List[Tuple[int, str]], List[Tuple[str, str]]]:
    """Envía las tarjetas de los empleados al servicio externo de RRHH."""
    base_url = RRHH_CARD_UPDATE_URL
    if not base_url:
        raise RuntimeError(
            "RRHH_CARD_UPDATE_URL no está configurada. Indica la URL en el archivo .env."
        )

    attempts = 0
    successes = 0
    success_entries: List[Tuple[int, str]] = []
    errors: List[Tuple[str, str]] = []

    for employee in employees:
        numeric_user_id = _normalize_numeric_identifier(
            employee.get("user_id") or employee.get("code") or employee.get("uid")
        )
        alias_value = str(numeric_user_id) if numeric_user_id is not None else None
        card_value = _normalize_positive_numeric_text(employee.get("tarjeta") or employee.get("card"))

        if numeric_user_id is None or alias_value is None or card_value is None:
            continue

        attempts += 1
        query_params = urlencode(
            {
                "c": str(numeric_user_id),
                "a": alias_value,
                "t": card_value,
            }
        )
        request_url = f"{base_url}?{query_params}"

        try:
            with urlopen(request_url, timeout=RRHH_CARD_UPDATE_TIMEOUT) as response:
                status_code = getattr(response, "status", None) or response.getcode()
                if status_code and status_code >= 400:
                    raise URLError(f"HTTP {status_code}")
                # Consumimos la respuesta para liberar la conexión si usa keep-alive
                response.read()
        except Exception as exc:  # pragma: no cover - dependiente del servicio externo
            logger.exception(
                "Error al actualizar la tarjeta en RRHH para el empleado %s: %s",
                numeric_user_id,
                exc,
            )
            errors.append((str(numeric_user_id), str(exc)))
        else:
            successes += 1
            success_entries.append((numeric_user_id, card_value))

    return successes, attempts, success_entries, errors


EXTERNAL_EMPLOYEE_URL = get_setting(
    "EXTERNAL_EMPLOYEE_URL", "http://lpa6.bonny.eu:8888/rh/zk.employees"
)
EXTERNAL_EMPLOYEE_CACHE_TTL = timedelta(hours=2)
EXTERNAL_EMPLOYEE_CACHE: Dict[str, object] = {"data": None, "timestamp": None}

EXPORT_COLUMNS: Sequence[Tuple[str, str]] = (
    ("uid", "UID"),
    ("name", "Nombre"),
    ("user_id", "User ID"),
    ("card", "Tarjeta"),
    ("privilege", "Privilegio"),
    ("group_id", "Grupo"),
    ("contract_from", "Contrato desde"),
    ("medical_leave_from", "IT desde"),
    ("vacation_status", "Vacaciones"),
    ("biometrics", "Biometría"),
)

SPANISH_MONTH_ABBR = {
    1: "Ene",
    2: "Feb",
    3: "Mar",
    4: "Abr",
    5: "May",
    6: "Jun",
    7: "Jul",
    8: "Ago",
    9: "Sep",
    10: "Oct",
    11: "Nov",
    12: "Dic",
}


def fetch_employees(host: str, port: int = DEFAULT_PORT) -> List[dict]:
    """Obtiene los empleados del terminal incluyendo datos biométricos."""
    zk = conn = None
    try:
        zk, conn = connect_with_retries(host, port)
        users = conn.get_users()
        try:
            templates = conn.get_templates()
        except Exception as exc:  # pragma: no cover - depende del terminal
            logger.warning("No fue posible obtener plantillas biométricas: %s", exc)
            templates = []

        template_index: Dict[int, List[dict]] = {}
        for template in templates or []:
            uid = getattr(template, "uid", None)
            if uid is None:
                continue
            template_index.setdefault(uid, []).append(
                {
                    "fid": getattr(template, "fid", ""),
                    "type": getattr(template, "type", ""),
                    "valid": getattr(template, "valid", ""),
                    "size": len(getattr(template, "template", b""))
                    if hasattr(template, "template")
                    else None,
                }
            )

        employees: List[dict] = []
        for user in users:
            uid = getattr(user, "uid", "")
            employees.append(
                {
                    "uid": str(uid),
                    "name": getattr(user, "name", ""),
                    "user_id": getattr(user, "user_id", ""),
                    "card": getattr(user, "card", ""),
                    "privilege": getattr(user, "privilege", ""),
                    "group_id": getattr(user, "group_id", ""),
                    "biometrics": template_index.get(uid, []),
                }
            )
        return employees
    finally:
        try:
            if conn:
                conn.disconnect()
        except Exception:  # pragma: no cover - errores de red
            logger.exception("Error al desconectar del terminal %s", host)


def delete_employees(
    host: str, employees: Iterable[dict], port: int = DEFAULT_PORT
) -> Tuple[List[str], List[Tuple[str, str]]]:
    """Elimina empleados del terminal y devuelve listas de eliminados y errores."""
    zk = conn = None
    deleted: List[str] = []
    errors: List[Tuple[str, str]] = []
    try:
        zk, conn = connect_with_retries(host, port)
        for employee in employees:
            uid_str = employee.get("uid", "")
            kwargs = {}
            try:
                uid_value = int(uid_str)
            except (TypeError, ValueError):
                uid_value = None

            if uid_value is not None:
                kwargs["uid"] = uid_value

            user_id = (employee.get("user_id") or "").strip()
            if user_id:
                kwargs["user_id"] = user_id

            if not kwargs:
                errors.append((str(uid_str), "El registro no tiene identificadores válidos"))
                continue

            try:
                conn.delete_user(**kwargs)
            except Exception as exc:  # pragma: no cover - depende del terminal
                errors.append((str(uid_str), str(exc)))
            else:
                deleted.append(str(uid_str))
    finally:
        try:
            if conn:
                conn.disconnect()
        except Exception:  # pragma: no cover - errores de red
            logger.exception("Error al desconectar del terminal %s", host)

    return deleted, errors


def get_terminal_status(host: str, port: int = DEFAULT_PORT) -> Tuple[dict, List[str]]:
    """Recopila información general del terminal para mostrar al usuario."""

    zk = conn = None
    info: Dict[str, Optional[str]] = {
        "Dirección IP": host,
        "Puerto": str(port),
    }
    errors: List[str] = []

    def safe_call(attr: str, label: str) -> Optional[str]:
        if not hasattr(conn, attr):
            errors.append(f"El terminal no soporta la consulta '{label}'.")
            return None
        try:
            value = getattr(conn, attr)()
        except Exception as exc:  # pragma: no cover - dependiente del dispositivo
            errors.append(f"Error al obtener {label}: {exc}")
            return None
        if value is None:
            return None
        return str(value)

    try:
        zk, conn = connect_with_retries(host, port)
        if conn is None:
            raise ValueError("No se pudo establecer conexión con el terminal.")

        info["Número de serie"] = safe_call("get_serialnumber", "el número de serie")
        info["Nombre del dispositivo"] = safe_call("get_device_name", "el nombre del dispositivo")
        info["Modelo"] = safe_call("get_model", "el modelo")
        info["Plataforma"] = safe_call("get_platform", "la plataforma")
        info["Versión de firmware"] = safe_call("get_firmware_version", "la versión de firmware")
        info["Dirección MAC"] = safe_call("get_mac", "la dirección MAC")
        info["Fecha y hora"] = safe_call("get_time", "la fecha y hora")

        users: List = []
        try:
            users = conn.get_users() or []
        except Exception as exc:  # pragma: no cover - dependiente del dispositivo
            errors.append(f"No se pudo obtener la lista de usuarios: {exc}")
        else:
            info["Usuarios en memoria"] = str(len(users))

        attendance_count: Optional[int] = None
        if hasattr(conn, "get_attendance_count"):
            try:
                attendance_count = conn.get_attendance_count()
            except Exception as exc:  # pragma: no cover - dependiente del dispositivo
                errors.append(f"No se pudo obtener el número de marcajes: {exc}")
        elif hasattr(conn, "get_attendance"):
            try:
                attendances = conn.get_attendance() or []
            except Exception as exc:  # pragma: no cover - dependiente del dispositivo
                errors.append(f"No se pudo obtener los marcajes: {exc}")
            else:
                attendance_count = len(attendances)

        if attendance_count is not None:
            info["Marcajes en memoria"] = str(attendance_count)

        if hasattr(conn, "get_work_code"):
            try:
                workcodes = conn.get_work_code() or []
            except Exception as exc:  # pragma: no cover - dependiente del dispositivo
                errors.append(f"No se pudo obtener los códigos de trabajo: {exc}")
            else:
                info["Códigos de trabajo"] = str(len(workcodes))

    finally:
        try:
            if conn:
                conn.disconnect()
        except Exception:  # pragma: no cover - errores de red
            logger.exception("Error al desconectar del terminal %s", host)

    cleaned_info = {k: v for k, v in info.items() if v}
    return cleaned_info, errors


def load_known_terminals() -> List[Dict[str, str]]:
    """Devuelve la lista de terminales conocidos desde el fichero `terminales.txt`."""

    try:
        content = TERMINAL_LIST_PATH.read_text(encoding="utf-8")
    except FileNotFoundError:
        logger.warning("No se encontró el fichero de terminales en %s", TERMINAL_LIST_PATH)
        return []
    except OSError as exc:
        logger.warning("No se pudo leer el fichero de terminales: %s", exc)
        return []

    terminals: List[Dict[str, str]] = []
    seen_ips = set()
    for line in content.splitlines():
        entry = line.strip()
        if not entry or entry.startswith("#"):
            continue
        label = ""
        ip = entry
        if "," in entry:
            name_part, ip_part = entry.split(",", 1)
            label = name_part.strip()
            ip = ip_part.strip() or ip
        elif " " in entry:
            parts = entry.split()
            ip_candidate = parts[-1]
            ip = ip_candidate.strip()
            label = entry[: -len(ip_candidate)].strip()
        if not ip or ip in seen_ips:
            continue
        seen_ips.add(ip)
        terminals.append({
            "ip": ip,
            "label": label,
        })
    return terminals


def upload_employees(
    host: str, employees: Iterable[dict], port: int = DEFAULT_PORT
) -> Tuple[List[str], List[Tuple[str, str]]]:
    """Envía empleados al terminal creando o actualizando sus datos básicos."""

    allowed_privileges = {
        int(getattr(const, name))
        for name in ("USER_DEFAULT", "USER_ADMIN", "USER_ENROLLER", "USER_SUPERADMIN")
        if hasattr(const, name)
    }

    def _coerce_privilege(value) -> int:
        if value is None:
            return const.USER_DEFAULT
        if isinstance(value, int):
            return value
        text = str(value).strip()
        if not text:
            return const.USER_DEFAULT
        lowered = text.lower()
        mapping = {
            "admin": const.USER_ADMIN,
            "superadmin": const.USER_ADMIN,
            "user": const.USER_DEFAULT,
            "default": const.USER_DEFAULT,
        }
        if lowered in mapping:
            return mapping[lowered]
        try:
            numeric = int(text)
        except ValueError:
            return const.USER_DEFAULT
        if allowed_privileges and numeric not in allowed_privileges:
            return const.USER_DEFAULT
        return numeric

    zk = conn = None
    uploaded: List[str] = []
    errors: List[Tuple[str, str]] = []
    try:
        zk, conn = connect_with_retries(host, port)
        if conn is None:
            raise ValueError("No se pudo establecer conexión con el terminal.")
        try:
            conn.disable_device()
        except Exception as exc:  # pragma: no cover - dependiente del terminal
            logger.warning("No fue posible deshabilitar temporalmente el terminal %s: %s", host, exc)

        try:
            existing_users = conn.get_users()
        except Exception as exc:  # pragma: no cover - dependiente del terminal
            logger.warning("No fue posible recuperar usuarios existentes de %s: %s", host, exc)
            existing_users = []

        used_uids: Set[int] = set()
        for user in existing_users:
            try:
                uid_int = int(getattr(user, "uid", None))
            except (TypeError, ValueError):
                continue
            else:
                used_uids.add(uid_int)

        allocated_uids: Set[int] = set()
        next_candidate = 1

        def _allocate_uid() -> int:
            nonlocal next_candidate
            attempts = 0
            while True:
                if next_candidate not in used_uids and next_candidate not in allocated_uids:
                    allocated_uids.add(next_candidate)
                    assigned = next_candidate
                    next_candidate += 1
                    return assigned
                next_candidate += 1
                attempts += 1
                if attempts > 200000:  # límite razonable para evitar bucles infinitos
                    raise ValueError("No se encontraron UID libres en el terminal.")

        for employee in employees:
            uid_label = str(employee.get("uid", "") or "").strip()
            name = str(employee.get("name", "") or "").strip()
            user_id = str(employee.get("user_id", "") or "").strip()
            group_id = str(employee.get("group_id", "") or "").strip()
            privilege = _coerce_privilege(employee.get("privilege"))
            card_value = employee.get("card")
            card = str(card_value).strip() if card_value is not None else ""
            if card.lower() in {"", "0", "none", "null"}:
                card = None

            if not uid_label and not user_id:
                errors.append(("(sin identificador)", "UID o User ID inválido"))
                continue

            try:
                new_uid = _allocate_uid()
            except ValueError as exc:
                errors.append((uid_label or user_id or "(sin identificador)", str(exc)))
                break

            payload = {
                "uid": new_uid,
                "name": name,
                "privilege": privilege,
                "group_id": group_id or "",
                "user_id": user_id or "",
                "password": "",
            }
            if card is not None:
                payload["card"] = card

            try:
                conn.set_user(**payload)
            except Exception as exc:  # pragma: no cover - dependiente del terminal
                errors.append((uid_label or user_id or "(sin UID)", str(exc)))
            else:
                uploaded.append(uid_label or str(new_uid))
    finally:
        try:
            if conn:
                try:
                    conn.enable_device()
                except Exception:  # pragma: no cover - dependiente del terminal
                    logger.exception("Error al habilitar nuevamente el terminal %s", host)
                conn.disconnect()
        except Exception:  # pragma: no cover - errores de red
            logger.exception("Error al desconectar del terminal %s", host)

    return uploaded, errors


def sync_terminal_time(host: str, port: int = DEFAULT_PORT) -> None:
    """Sincroniza la fecha y hora del terminal con la del sistema."""

    zk = conn = None
    try:
        zk, conn = connect_with_retries(host, port)
        if conn is None:
            raise ValueError("No se pudo establecer conexión con el terminal.")
        conn.enable_device()
        conn.set_time(datetime.now())
    finally:
        try:
            if conn:
                conn.disconnect()
        except Exception:  # pragma: no cover - errores de red
            logger.exception("Error al desconectar del terminal %s", host)


def _stringify_export_value(value):
    if value is None:
        return ""
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def build_export_response(host: str, employees: List[dict], export_format: str):
    """Genera un archivo de exportación para los empleados seleccionados."""
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    safe_host = host.replace(":", "-")
    base_filename = f"empleados_{safe_host}_{timestamp}"

    if export_format == "json":
        payload = json.dumps(employees, ensure_ascii=False, indent=2)
        response = make_response(payload)
        response.headers["Content-Type"] = "application/json; charset=utf-8"
        response.headers["Content-Disposition"] = (
            f"attachment; filename={base_filename}.json"
        )
        return response

    if export_format == "csv":
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow([header for _, header in EXPORT_COLUMNS])
        for employee in employees:
            row = [_stringify_export_value(employee.get(key)) for key, _ in EXPORT_COLUMNS]
            writer.writerow(row)
        csv_data = output.getvalue()
        response = make_response(csv_data)
        response.headers["Content-Type"] = "text/csv; charset=utf-8"
        response.headers["Content-Disposition"] = (
            f"attachment; filename={base_filename}.csv"
        )
        return response

    if export_format == "excel":
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Empleados"
        worksheet.append([header for _, header in EXPORT_COLUMNS])
        for employee in employees:
            row = [_stringify_export_value(employee.get(key)) for key, _ in EXPORT_COLUMNS]
            worksheet.append(row)
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name=f"{base_filename}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    raise ValueError("Formato de exportación no soportado.")


def coerce_port(port_value: Optional[str]) -> int:
    try:
        port = int(port_value) if port_value is not None else DEFAULT_PORT
    except (TypeError, ValueError):
        return DEFAULT_PORT
    if 1 <= port <= 65535:
        return port
    return DEFAULT_PORT


def parse_terminal_value(value: Optional[str]) -> Tuple[Optional[str], int]:
    """Convierte el texto introducido por el usuario en IP y puerto."""

    if not value:
        return None, DEFAULT_PORT

    cleaned = value.strip()
    if not cleaned:
        return None, DEFAULT_PORT

    host = cleaned
    port = DEFAULT_PORT

    if cleaned.startswith("[") and "]" in cleaned:
        closing = cleaned.find("]")
        host = cleaned[1:closing].strip()
        remainder = cleaned[closing + 1 :].strip()
        if remainder.startswith(":"):
            port = coerce_port(remainder[1:].strip())
    else:
        parts = cleaned.rsplit(":", 1)
        if len(parts) == 2 and parts[0]:
            potential_host, potential_port = parts
            if potential_port:
                host = potential_host.strip() or cleaned
                port = coerce_port(potential_port.strip())
        elif cleaned.count(":") > 1:
            host = cleaned  # Dirección IPv6 sin puerto

    host = host.strip()
    if not host:
        host = None

    return host, port


def format_terminal_value(host: Optional[str], port: int) -> str:
    if not host:
        return ""
    if port == DEFAULT_PORT:
        return host
    return f"{host}:{port}"


def format_contract_date(value: Optional[str]) -> str:
    """Convierte una fecha ISO en formato ``ddMonyy`` con meses en español."""
    if not value:
        return ""

    text = str(value).strip()
    if not text:
        return ""

    candidate = text.replace("Z", "+00:00")

    parsed: Optional[datetime] = None
    for parser in (
        lambda val: datetime.fromisoformat(val),
        lambda val: datetime.strptime(val, "%Y-%m-%d"),
        lambda val: datetime.strptime(val, "%Y/%m/%d"),
    ):
        try:
            parsed = parser(candidate)
            break
        except ValueError:
            continue

    if parsed is None:
        return text

    # Normaliza a naive en UTC si tiene tz
    if parsed.tzinfo is not None:
        parsed = parsed.astimezone(timezone.utc).replace(tzinfo=None)

    month_abbr = SPANISH_MONTH_ABBR.get(parsed.month)
    if not month_abbr:
        return text

    return f"{parsed.day:02}{month_abbr}{parsed.year % 100:02}"


def _normalize_employee_record(raw: dict) -> dict:
    """Normaliza los datos de un empleado importado."""

    key_aliases = {
        "uid": {"uid", "id", "identificador"},
        "name": {"name", "nombre"},
        "user_id": {"user_id", "user id", "userid", "id usuario", "idusuario"},
        "card": {"card", "tarjeta", "num tarjeta"},
        "privilege": {"privilege", "privilegio"},
        "group_id": {"group_id", "group id", "grupo", "id grupo", "grupo id"},
        "contract_from": {"contrato_desde", "contrato desde", "contract_from"},
        "medical_leave_from": {"it_desde", "it desde", "medical_leave_from"},
        "vacation_status": {"vacaciones", "vacation_status"},
        "biometrics": {"biometrics", "biometria", "biometría", "biometricas", "plantillas"},
    }

    normalized_keys = {}
    for original_key, value in raw.items():
        if isinstance(original_key, str):
            normalized_keys[original_key.strip().lower()] = value

    def _safe_get(key: str):
        candidates = key_aliases.get(key, set()) | {key, key.lower(), key.upper()}
        for candidate in candidates:
            if not isinstance(candidate, str):
                continue
            candidate_key = candidate.strip().lower()
            if candidate_key in normalized_keys:
                return normalized_keys[candidate_key]
            if candidate in raw:
                return raw.get(candidate)
        return raw.get(key)

    biometrics_value = _safe_get("biometrics")
    biometrics: List[dict]
    if isinstance(biometrics_value, list):
        biometrics = [item for item in biometrics_value if isinstance(item, dict)]
    elif isinstance(biometrics_value, str):
        biometrics_value = biometrics_value.strip()
        if biometrics_value:
            try:
                parsed = json.loads(biometrics_value)
            except json.JSONDecodeError:
                biometrics = []
            else:
                biometrics = (
                    [item for item in parsed if isinstance(item, dict)]
                    if isinstance(parsed, list)
                    else []
                )
        else:
            biometrics = []
    else:
        biometrics = []

    return {
        "uid": str(_safe_get("uid") or ""),
        "name": _safe_get("name") or "",
        "user_id": _safe_get("user_id") or "",
        "card": _safe_get("card") or "",
        "privilege": _safe_get("privilege") or "",
        "group_id": _safe_get("group_id") or "",
        "contract_from": _safe_get("contract_from") or "",
        "medical_leave_from": _safe_get("medical_leave_from") or "",
        "vacation_status": _safe_get("vacation_status") or "",
        "biometrics": biometrics,
    }


def parse_employee_file(file_storage) -> List[dict]:
    """Convierte un archivo cargado en una lista de empleados."""

    if file_storage is None:
        raise ValueError("Selecciona un archivo para importar.")

    filename = (file_storage.filename or "").strip()
    if not filename:
        raise ValueError("Selecciona un archivo para importar.")

    payload = file_storage.read()
    if not payload:
        raise ValueError("El archivo de empleados está vacío.")

    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext == "json":
        try:
            data = json.loads(payload.decode("utf-8-sig"))
        except json.JSONDecodeError as exc:
            raise ValueError(f"El archivo JSON es inválido: {exc}") from exc
        if not isinstance(data, list):
            raise ValueError("El archivo JSON debe contener una lista de empleados.")
        return [_normalize_employee_record(item) for item in data if isinstance(item, dict)]

    if ext == "csv":
        text = payload.decode("utf-8-sig")
        reader = csv.DictReader(StringIO(text))
        return [_normalize_employee_record({k.lower(): v for k, v in row.items()}) for row in reader]

    if ext in {"xlsx", "xlsm"}:
        workbook = load_workbook(BytesIO(payload), data_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
        employees: List[dict] = []
        for row in rows[1:]:
            if row is None:
                continue
            row_dict = {
                headers[index]: value
                for index, value in enumerate(row)
                if index < len(headers) and headers[index]
            }
            employees.append(_normalize_employee_record(row_dict))
        return employees

    raise ValueError("Formato de archivo no soportado. Usa JSON, CSV o Excel.")


def _normalize_database_employee_record(raw: dict) -> Optional[dict]:
    """Convierte un registro externo en el esquema usado por la aplicación."""
    if not isinstance(raw, dict):
        return None

    def _first_non_empty(values):
        for candidate in values:
            text = str(candidate or "").strip()
            if text:
                return text
        return ""

    user_code = _first_non_empty(
        [
            raw.get("CODIGO_ZK_ATRIBUTO"),
            raw.get("user_id"),
            raw.get("uid"),
            raw.get("UID"),
            raw.get("id"),
            raw.get("ID"),
            raw.get("DNI"),
        ]
    )
    if not user_code:
        return None

    name = _first_non_empty([
        raw.get("NOMBRE"),
        raw.get("nombre"),
        raw.get("name"),
        raw.get("NAME"),
    ])

    card = _first_non_empty([
        raw.get("card"),
        raw.get("CARD"),
        raw.get("CARDNUMBER"),
        raw.get("NUM_TARJETA"),
        raw.get("NUMERO_TARJETA"),
    ])

    privilege = _first_non_empty([
        raw.get("privilege"),
        raw.get("PRIVILEGIO"),
        raw.get("privilegio"),
    ])

    group_id = _first_non_empty([
        raw.get("group_id"),
        raw.get("GROUP_ID"),
        raw.get("COD_CT"),
        raw.get("cod_ct"),
    ])

    dni = _first_non_empty([
        raw.get("DNI"),
        raw.get("dni"),
    ])

    last_seen = raw.get("LAST_SEEN") or raw.get("last_seen")
    contract_from = raw.get("CONTRATO_DESDE") or raw.get("contrato_desde")
    medical_leave_from = raw.get("IT_DESDE") or raw.get("it_desde")
    vacation_status = raw.get("VACACIONES") or raw.get("vacaciones")

    return {
        "uid": user_code,
        "name": name or user_code,
        "user_id": user_code,
        "card": card,
        "privilege": privilege,
        "group_id": group_id,
        "biometrics": [],
        "dni": dni,
        "last_seen": last_seen,
        "contract_from": contract_from,
        "medical_leave_from": medical_leave_from,
        "vacation_status": vacation_status,
    }


def get_cached_employees(host: str) -> List[dict]:
    """Devuelve los empleados almacenados en memoria para un terminal."""
    return TERMINAL_EMPLOYEES.get(host, [])


def refresh_database_cache() -> List[dict]:
    """Refresca la caché de empleados externos y devuelve los registros normalizados."""
    try:
        data = load_external_employees(force_refresh=True)
    except Exception:
        # Propaga el error para que la vista decida cómo manejarlo.
        raise

    normalized: List[dict] = []
    for record in data:
        normalized_record = _normalize_database_employee_record(record)
        if normalized_record is None:
            continue
        normalized.append(normalized_record)

    TERMINAL_EMPLOYEES[DATABASE_TERMINAL_KEY] = normalized
    SELECTED_EMPLOYEES.pop(DATABASE_TERMINAL_KEY, None)
    return normalized


def set_cached_employees(host: str, employees: List[dict]) -> None:
    """Guarda los empleados en memoria para un terminal."""
    TERMINAL_EMPLOYEES[host] = employees


def clear_terminal_cache(host: str) -> List[dict]:
    """Elimina y devuelve los empleados en memoria de un terminal."""
    removed = TERMINAL_EMPLOYEES.pop(host, [])
    SELECTED_EMPLOYEES.pop(host, None)
    return removed


def clear_all_cache() -> None:
    """Vacía las estructuras en memoria utilizadas por la aplicación."""
    TERMINAL_EMPLOYEES.clear()
    SELECTED_EMPLOYEES.clear()


def get_selected_uids(host: str) -> Set[str]:
    """Obtiene los UID seleccionados para un terminal."""
    return set(SELECTED_EMPLOYEES.get(host, set()))


def set_selected_uids(host: str, selected: Iterable[str]) -> None:
    """Almacena los UID seleccionados para un terminal."""
    SELECTED_EMPLOYEES[host] = set(selected)


def remove_selected_uids(host: str, uids: Iterable[str]) -> None:
    """Elimina UID concretos del conjunto de seleccionados de un terminal."""
    existing = SELECTED_EMPLOYEES.get(host)
    if existing is None:
        return
    existing.difference_update({str(uid) for uid in uids})
    SELECTED_EMPLOYEES[host] = existing


def find_duplicate_employees(employees: Iterable[dict]) -> List[dict]:
    """Devuelve empleados que comparten nombre pero tienen distinto user_id."""
    name_to_user_ids: Dict[str, Set[str]] = defaultdict(set)

    normalized_employees: List[Tuple[str, dict]] = []
    for employee in employees:
        raw_name = (employee.get("name") or "").strip()
        normalized_name = raw_name.casefold()
        user_id_value = str(employee.get("user_id") or "").strip()

        normalized_employees.append((normalized_name, employee))
        name_to_user_ids[normalized_name].add(user_id_value)

    duplicate_names = {
        normalized_name
        for normalized_name, user_ids in name_to_user_ids.items()
        if len(user_ids) > 1 and normalized_name
    }

    if not duplicate_names:
        return []

    return [
        employee
        for normalized_name, employee in normalized_employees
        if normalized_name in duplicate_names
    ]


def _download_external_employees() -> List[dict]:
    """Descarga la lista de empleados externos desde el endpoint remoto."""
    try:
        with urlopen(EXTERNAL_EMPLOYEE_URL, timeout=15) as response:  # nosec: B310 - endpoint controlado
            status = getattr(response, "status", response.getcode())
            if status and status >= 400:
                raise ValueError(f"Respuesta inesperada ({status}) al consultar empleados externos.")
            payload = response.read()
    except URLError as exc:
        raise RuntimeError(f"No se pudo conectar con la fuente de empleados externa: {exc}") from exc

    try:
        data = json.loads(payload.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise ValueError("La respuesta de empleados externa no es un JSON válido.") from exc

    if not isinstance(data, list):
        raise ValueError("La respuesta de empleados externa no es una lista.")

    return data


def load_external_employees(force_refresh: bool = False) -> List[dict]:
    """Obtiene la lista de empleados externos, usando caché durante 2 horas."""
    cached_data = EXTERNAL_EMPLOYEE_CACHE.get("data")
    cached_timestamp = EXTERNAL_EMPLOYEE_CACHE.get("timestamp")
    now = datetime.now(timezone.utc)

    if (
        not force_refresh
        and cached_data
        and isinstance(cached_timestamp, datetime)
        and now - cached_timestamp < EXTERNAL_EMPLOYEE_CACHE_TTL
    ):
        return cached_data

    try:
        data = _download_external_employees()
    except Exception as exc:
        logger.exception("Error al refrescar los empleados externos: %s", exc)
        if cached_data:
            return cached_data
        raise

    EXTERNAL_EMPLOYEE_CACHE["data"] = data
    EXTERNAL_EMPLOYEE_CACHE["timestamp"] = now
    return data


def _store_external_mapping_entry(
    mapping: Dict[str, dict],
    user_code: str,
    payload: dict,
) -> None:
    """Agrega un registro al mapeo usando variaciones comunes del código."""
    if not user_code:
        return

    mapping[user_code] = payload

    trimmed = user_code.lstrip("0")
    if trimmed and trimmed != user_code:
        mapping.setdefault(trimmed, payload)

    upper_original = user_code.upper()
    if upper_original not in mapping:
        mapping[upper_original] = payload

    if trimmed:
        upper_trimmed = trimmed.upper()
        if upper_trimmed not in mapping:
            mapping[upper_trimmed] = payload


def get_external_employee_map(force_refresh: bool = False) -> Dict[str, dict]:
    """Devuelve un diccionario user_id -> datos ampliados del empleado."""
    external_employees = load_external_employees(force_refresh=force_refresh)
    mapping: Dict[str, dict] = {}
    for record in external_employees:
        user_code = str(record.get("CODIGO_ZK_ATRIBUTO") or "").strip()
        if not user_code:
            continue
        payload = {
            "dni": str(record.get("DNI") or "").strip(),
            "nombre": str(record.get("NOMBRE") or "").strip(),
            "cod_ct": str(record.get("COD_CT") or "").strip(),
            "last_seen": record.get("LAST_SEEN"),
            "contract_from": _normalize_setting(
                record.get("CONTRATO_DESDE") or record.get("contrato_desde")
            ),
            "medical_leave_from": _normalize_setting(
                record.get("IT_DESDE") or record.get("it_desde")
            ),
            "vacation_status": _normalize_setting(
                record.get("VACACIONES") or record.get("vacaciones")
            ),
        }
        _store_external_mapping_entry(mapping, user_code, payload)
    return mapping


def get_external_employee_map_by_dni(force_refresh: bool = False) -> Dict[str, dict]:
    """Construye un diccionario con clave DNI para datos ampliados."""
    external_employees = load_external_employees(force_refresh=force_refresh)
    mapping: Dict[str, dict] = {}
    for record in external_employees:
        dni = str(record.get("DNI") or "").strip()
        if not dni:
            continue
        payload = {
            "dni": dni,
            "nombre": str(record.get("NOMBRE") or "").strip(),
            "cod_ct": str(record.get("COD_CT") or "").strip(),
            "last_seen": record.get("LAST_SEEN"),
            "contract_from": _normalize_setting(
                record.get("CONTRATO_DESDE") or record.get("contrato_desde")
            ),
            "medical_leave_from": _normalize_setting(
                record.get("IT_DESDE") or record.get("it_desde")
            ),
            "vacation_status": _normalize_setting(
                record.get("VACACIONES") or record.get("vacaciones")
            ),
        }
        normalized = dni.upper()
        mapping.setdefault(dni, payload)
        mapping.setdefault(normalized, payload)
    return mapping


def lookup_external_employee(identifier: str, mapping: Dict[str, dict]) -> Optional[dict]:
    """Busca un empleado externo usando variaciones comunes del identificador."""
    if not identifier:
        return None

    candidate = str(identifier).strip()
    if not candidate:
        return None

    variations = [
        candidate,
        candidate.upper(),
    ]

    trimmed = candidate.lstrip("0")
    if trimmed:
        variations.extend([trimmed, trimmed.upper()])

    for key in variations:
        if key and key in mapping:
            return mapping[key]
    return None


def format_relative_time(value: Optional[str], default: str = "N/A") -> str:
    """Convierte una fecha ISO en un texto relativo (p.ej. '4 hours ago')."""
    if not value:
        return default

    try:
        parsed = datetime.fromisoformat(value)
    except ValueError:
        try:
            parsed_date = datetime.strptime(value, "%Y-%m-%d")
        except ValueError:
            return default
        parsed = datetime(
            parsed_date.year,
            parsed_date.month,
            parsed_date.day,
            tzinfo=timezone.utc,
        )

    if not isinstance(parsed, datetime):
        parsed = datetime(parsed.year, parsed.month, parsed.day, tzinfo=timezone.utc)

    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)

    now = datetime.now(timezone.utc)
    delta = now - parsed
    seconds = delta.total_seconds()

    if abs(seconds) < 60:
        return "just now"

    future = seconds < 0
    seconds = abs(seconds)

    units = (
        (365 * 24 * 3600, "year"),
        (30 * 24 * 3600, "month"),
        (7 * 24 * 3600, "week"),
        (24 * 3600, "day"),
        (3600, "hour"),
        (60, "minute"),
    )

    for unit_seconds, unit_name in units:
        if seconds >= unit_seconds:
            value_count = int(seconds // unit_seconds)
            plural = "s" if value_count != 1 else ""
            text = f"{value_count} {unit_name}{plural}"
            return f"in {text}" if future else f"{text} ago"

    return default
