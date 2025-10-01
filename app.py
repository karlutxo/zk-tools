"""Aplicación Flask para consultar y seleccionar empleados de terminales ZKTeco."""
from __future__ import annotations

import csv
import json
import logging
from datetime import datetime
from io import BytesIO, StringIO
from typing import Dict, List, Optional, Sequence, Set, Tuple

from flask import (
    Flask,
    flash,
    make_response,
    redirect,
    render_template_string,
    request,
    send_file,
    url_for,
)
from openpyxl import Workbook, load_workbook

from zk_tools import DEFAULT_PORT, connect_with_retries

app = Flask(__name__)
app.secret_key = "zk-tools-dev"

# Almacenamiento en memoria de los empleados descubiertos por terminal
_TERMINAL_EMPLOYEES: Dict[str, List[dict]] = {}
# Almacenamiento en memoria de los UID seleccionados por terminal
_SELECTED_EMPLOYEES: Dict[str, Set[str]] = {}

EXPORT_COLUMNS: Sequence[Tuple[str, str]] = (
    ("uid", "UID"),
    ("name", "Nombre"),
    ("user_id", "User ID"),
    ("card", "Tarjeta"),
    ("privilege", "Privilegio"),
    ("group_id", "Grupo"),
    ("biometrics", "Biometría"),
)

logger = logging.getLogger(__name__)


def fetch_employees(host: str, port: int = DEFAULT_PORT) -> List[dict]:
    """Obtiene los empleados del terminal incluyendo datos biométricos."""
    zk = conn = None
    try:
        zk, conn = connect_with_retries(host, port)
        users = conn.get_users()
        try:
            templates = conn.get_templates()
        except Exception as exc:  # pragma: no cover - depende del terminal
            logger.warning("No fue posible obtener plantillas biométricas: %s", exc)
            templates = []

        template_index: Dict[int, List[dict]] = {}
        for template in templates or []:
            uid = getattr(template, "uid", None)
            if uid is None:
                continue
            template_index.setdefault(uid, []).append(
                {
                    "fid": getattr(template, "fid", ""),
                    "type": getattr(template, "type", ""),
                    "valid": getattr(template, "valid", ""),
                    "size": len(getattr(template, "template", b"")) if hasattr(template, "template") else None,
                }
            )

        employees: List[dict] = []
        for user in users:
            uid = getattr(user, "uid", "")
            employees.append(
                {
                    "uid": str(uid),
                    "name": getattr(user, "name", ""),
                    "user_id": getattr(user, "user_id", ""),
                    "card": getattr(user, "card", ""),
                    "privilege": getattr(user, "privilege", ""),
                    "group_id": getattr(user, "group_id", ""),
                    "biometrics": template_index.get(uid, []),
                }
            )
        return employees
    finally:
        try:
            if conn:
                conn.disconnect()
        except Exception:  # pragma: no cover - errores de red
            logger.exception("Error al desconectar del terminal %s", host)


def delete_employees(
    host: str,
    employees: List[dict],
    port: int = DEFAULT_PORT,
) -> Tuple[List[str], List[Tuple[str, str]]]:
    """Elimina empleados del terminal y devuelve listas de eliminados y errores."""
    zk = conn = None
    deleted: List[str] = []
    errors: List[Tuple[str, str]] = []
    try:
        zk, conn = connect_with_retries(host, port)
        for employee in employees:
            uid_str = employee.get("uid", "")
            kwargs = {}
            try:
                uid_value = int(uid_str)
            except (TypeError, ValueError):
                uid_value = None

            if uid_value is not None:
                kwargs["uid"] = uid_value

            user_id = (employee.get("user_id") or "").strip()
            if user_id:
                kwargs["user_id"] = user_id

            if not kwargs:
                errors.append((str(uid_str), "El registro no tiene identificadores válidos"))
                continue

            try:
                conn.delete_user(**kwargs)
            except Exception as exc:  # pragma: no cover - depende del terminal
                errors.append((str(uid_str), str(exc)))
            else:
                deleted.append(str(uid_str))
    finally:
        try:
            if conn:
                conn.disconnect()
        except Exception:  # pragma: no cover - errores de red
            logger.exception("Error al desconectar del terminal %s", host)

    return deleted, errors


def _stringify_export_value(value):
    if value is None:
        return ""
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def build_export_response(host: str, employees: List[dict], export_format: str):
    """Genera un archivo de exportación para los empleados seleccionados."""
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    safe_host = host.replace(":", "-")
    base_filename = f"empleados_{safe_host}_{timestamp}"

    if export_format == "json":
        payload = json.dumps(employees, ensure_ascii=False, indent=2)
        response = make_response(payload)
        response.headers["Content-Type"] = "application/json; charset=utf-8"
        response.headers["Content-Disposition"] = (
            f"attachment; filename={base_filename}.json"
        )
        return response

    if export_format == "csv":
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow([header for _, header in EXPORT_COLUMNS])
        for employee in employees:
            row = [_stringify_export_value(employee.get(key)) for key, _ in EXPORT_COLUMNS]
            writer.writerow(row)
        csv_data = output.getvalue()
        response = make_response(csv_data)
        response.headers["Content-Type"] = "text/csv; charset=utf-8"
        response.headers["Content-Disposition"] = (
            f"attachment; filename={base_filename}.csv"
        )
        return response

    if export_format == "excel":
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Empleados"
        worksheet.append([header for _, header in EXPORT_COLUMNS])
        for employee in employees:
            row = [_stringify_export_value(employee.get(key)) for key, _ in EXPORT_COLUMNS]
            worksheet.append(row)
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name=f"{base_filename}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    raise ValueError("Formato de exportación no soportado.")


def _coerce_port(port_value: Optional[str]) -> int:
    try:
        port = int(port_value) if port_value is not None else DEFAULT_PORT
    except (TypeError, ValueError):
        return DEFAULT_PORT
    if 1 <= port <= 65535:
        return port
    return DEFAULT_PORT


def parse_terminal_value(value: Optional[str]) -> Tuple[Optional[str], int]:
    """Convierte el texto introducido por el usuario en IP y puerto."""

    if not value:
        return None, DEFAULT_PORT

    cleaned = value.strip()
    if not cleaned:
        return None, DEFAULT_PORT

    host = cleaned
    port = DEFAULT_PORT

    if cleaned.startswith("[") and "]" in cleaned:
        closing = cleaned.find("]")
        host = cleaned[1:closing].strip()
        remainder = cleaned[closing + 1 :].strip()
        if remainder.startswith(":"):
            port = _coerce_port(remainder[1:].strip())
    else:
        parts = cleaned.rsplit(":", 1)
        if len(parts) == 2 and parts[0]:
            potential_host, potential_port = parts
            if potential_port:
                host = potential_host.strip() or cleaned
                port = _coerce_port(potential_port.strip())
        elif cleaned.count(":") > 1:
            host = cleaned  # Dirección IPv6 sin puerto

    host = host.strip()
    if not host:
        host = None

    return host, port


def format_terminal_value(host: Optional[str], port: int) -> str:
    if not host:
        return ""
    if port == DEFAULT_PORT:
        return host
    return f"{host}:{port}"


def _normalize_employee_record(raw: dict) -> dict:
    """Normaliza los datos de un empleado importado."""

    def _safe_get(key: str):
        for candidate in (key, key.lower(), key.upper()):
            if candidate in raw:
                return raw.get(candidate)
        return raw.get(key)

    biometrics_value = _safe_get("biometrics")
    biometrics: List[dict]
    if isinstance(biometrics_value, list):
        biometrics = [item for item in biometrics_value if isinstance(item, dict)]
    elif isinstance(biometrics_value, str):
        biometrics_value = biometrics_value.strip()
        if biometrics_value:
            try:
                parsed = json.loads(biometrics_value)
            except json.JSONDecodeError:
                biometrics = []
            else:
                biometrics = (
                    [item for item in parsed if isinstance(item, dict)]
                    if isinstance(parsed, list)
                    else []
                )
        else:
            biometrics = []
    else:
        biometrics = []

    return {
        "uid": str(_safe_get("uid") or ""),
        "name": _safe_get("name") or "",
        "user_id": _safe_get("user_id") or "",
        "card": _safe_get("card") or "",
        "privilege": _safe_get("privilege") or "",
        "group_id": _safe_get("group_id") or "",
        "biometrics": biometrics,
    }


def parse_employee_file(file_storage) -> List[dict]:
    """Convierte un archivo cargado en una lista de empleados."""

    if file_storage is None:
        raise ValueError("Selecciona un archivo para importar.")

    filename = (file_storage.filename or "").strip()
    if not filename:
        raise ValueError("Selecciona un archivo para importar.")

    payload = file_storage.read()
    if not payload:
        raise ValueError("El archivo de empleados está vacío.")

    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext == "json":
        try:
            data = json.loads(payload.decode("utf-8-sig"))
        except json.JSONDecodeError as exc:
            raise ValueError(f"El archivo JSON es inválido: {exc}") from exc
        if not isinstance(data, list):
            raise ValueError("El archivo JSON debe contener una lista de empleados.")
        return [_normalize_employee_record(item) for item in data if isinstance(item, dict)]

    if ext == "csv":
        text = payload.decode("utf-8-sig")
        reader = csv.DictReader(StringIO(text))
        return [_normalize_employee_record({k.lower(): v for k, v in row.items()}) for row in reader]

    if ext in {"xlsx", "xlsm"}:
        workbook = load_workbook(BytesIO(payload), data_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
        employees: List[dict] = []
        for row in rows[1:]:
            if row is None:
                continue
            row_dict = {
                headers[index]: value
                for index, value in enumerate(row)
                if index < len(headers) and headers[index]
            }
            employees.append(_normalize_employee_record(row_dict))
        return employees

    raise ValueError("Formato de archivo no soportado. Usa JSON, CSV o Excel.")


INDEX_TEMPLATE = """
<!doctype html>
<html lang="es">
<head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>Empleados del terminal</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
    <link href="https://cdn.datatables.net/1.13.6/css/dataTables.bootstrap5.min.css" rel="stylesheet">
    <style>
        body { background-color: #f8f9fa; }
        main { padding-bottom: 4rem; }
        .selected-row { background-color: #e8f4ff !important; }
        .biometric-list { margin-bottom: 0; padding-left: 1.25rem; }
        .accordion-button:not(.collapsed) { background-color: #0d6efd; color: #fff; }
        .accordion-button:focus { box-shadow: none; }
        .dataTables_wrapper .dataTables_paginate .paginate_button { border-radius: 0.375rem; }
        .dataTables_wrapper .dataTables_paginate .paginate_button.current { background-color: #0d6efd !important; color: #fff !important; }
        .dataTables_wrapper .dataTables_length select { width: auto; }
    </style>
</head>
<body>
    <nav class="navbar navbar-expand-lg navbar-dark bg-primary">
        <div class="container-fluid">
            <span class="navbar-brand mb-0 h1">Gestión de empleados ZKTeco</span>
            <span id="selection-info" class="badge bg-light text-dark fs-6" data-total="{{ total_employees }}" data-selected="{{ selected_count }}" title="Resumen de la selección actual">Seleccionados: {{ selected_count }} de {{ total_employees }}</span>
        </div>
    </nav>
    <main class="container my-4">
        <form method="post" action="{{ url_for('index') }}" enctype="multipart/form-data" class="accordion mb-4 shadow-sm" id="controlAccordion">
            <div class="accordion-item">
                <h2 class="accordion-header" id="headingConnection">
                    <button class="accordion-button" type="button" data-bs-toggle="collapse" data-bs-target="#collapseConnection" aria-expanded="true" aria-controls="collapseConnection">
                        Terminales
                    </button>
                </h2>
                <div id="collapseConnection" class="accordion-collapse collapse show" aria-labelledby="headingConnection" data-bs-parent="#controlAccordion">
                    <div class="accordion-body">
                        <div class="row g-3 align-items-end">
                            <div class="col-12">
                                <label for="terminal" class="form-label">Terminal</label>
                                <input type="text" name="terminal" id="terminal" class="form-control" value="{{ terminal }}" placeholder="Ej. 192.168.1.10 o 192.168.1.10:4370" aria-describedby="terminalHelp">
                                <div id="terminalHelp" class="form-text">Introduce la dirección del terminal. Puedes añadir el puerto con el formato IP:PUERTO.</div>
                            </div>
                            <div class="col-12">
                                <div class="alert alert-info mb-0" role="status">
                                    Usa estas opciones para conectarte con el terminal. Las herramientas de importación y exportación se encuentran en la sección siguiente.
                                </div>
                            </div>
                            <div class="col-12 d-flex flex-wrap gap-2">
                                <button type="submit" name="action" value="fetch" class="btn btn-primary">Leer empleados</button>
                                <button type="submit" name="action" value="clear" class="btn btn-outline-secondary" onclick="return confirm('Esto eliminará a todos los empleados almacenados en memoria para este terminal. ¿Continuar?');">Limpiar</button>
                                <button type="submit" name="action" value="delete" form="employees-form" class="btn btn-danger" {% if not employees %}disabled{% endif %} onclick="return confirm('¿Eliminar empleados seleccionados del terminal?');">Eliminar seleccionados</button>
                            </div>
                        </div>
                    </div>
                </div>
            </div>
            <div class="accordion-item">
                <h2 class="accordion-header" id="headingActions">
                    <button class="accordion-button collapsed" type="button" data-bs-toggle="collapse" data-bs-target="#collapseActions" aria-expanded="false" aria-controls="collapseActions">
                        Exportación e Importación
                    </button>
                </h2>
                <div id="collapseActions" class="accordion-collapse collapse" aria-labelledby="headingActions" data-bs-parent="#controlAccordion">
                    <div class="accordion-body">
                        <div class="row g-3 align-items-end">
                            <div class="col-12 col-lg-6">
                                <label for="employee-file" class="form-label">Archivo de empleados</label>
                                <div class="input-group">
                                    <input type="file" name="employee_file" id="employee-file" class="form-control" accept=".json,.csv,.xlsx,.xlsm" aria-describedby="fileHelp">
                                    <button type="submit" name="action" value="import" class="btn btn-success" {% if not terminal %}disabled{% endif %}>Importar</button>
                                </div>
                                <div id="fileHelp" class="form-text">Selecciona un archivo JSON, CSV o Excel para cargar empleados en la aplicación.</div>
                            </div>
                            <div class="col-12 d-flex flex-wrap gap-2 justify-content-md-end">
                                <button type="submit" name="action" value="select" form="employees-form" class="btn btn-outline-primary" {% if not employees %}disabled{% endif %}>Guardar selección</button>
                                <div class="btn-group">
                                    <button type="button" class="btn btn-success dropdown-toggle" data-bs-toggle="dropdown" aria-expanded="false" {% if not employees %}disabled{% endif %}>
                                        Exportar selección
                                    </button>
                                    <ul class="dropdown-menu dropdown-menu-end">
                                        <li><button class="dropdown-item" type="submit" name="action" value="export_csv" form="employees-form" {% if not employees %}disabled{% endif %}>Exportar CSV</button></li>
                                        <li><button class="dropdown-item" type="submit" name="action" value="export_json" form="employees-form" {% if not employees %}disabled{% endif %}>Exportar JSON</button></li>
                                        <li><button class="dropdown-item" type="submit" name="action" value="export_excel" form="employees-form" {% if not employees %}disabled{% endif %}>Exportar Excel</button></li>
                                    </ul>
                                </div>
                            </div>
                        </div>
                        <p class="text-muted small mt-3 mb-0">Las acciones se aplican únicamente a los empleados seleccionados en la tabla de abajo.</p>
                    </div>
                </div>
            </div>
        </form>

        {% with messages = get_flashed_messages() %}
            {% if messages %}
                <div class="alert alert-warning alert-dismissible fade show" role="alert">
                    <ul class="mb-0">
                        {% for message in messages %}
                        <li>{{ message }}</li>
                        {% endfor %}
                    </ul>
                    <button type="button" class="btn-close" data-bs-dismiss="alert" aria-label="Cerrar"></button>
                </div>
            {% endif %}
        {% endwith %}

        {% if employees %}
        <div class="card shadow-sm">
            <div class="card-body">
                <form method="post" action="{{ url_for('index') }}" id="employees-form" class="table-responsive">
                    <input type="hidden" name="terminal" value="{{ terminal }}">
                    <table class="table table-striped table-hover align-middle" id="employees-table">
                        <thead class="table-light">
                            <tr>
                                <th scope="col">Seleccionar</th>
                                <th scope="col">UID</th>
                                <th scope="col">Nombre</th>
                                <th scope="col">User ID</th>
                                <th scope="col">Tarjeta</th>
                                <th scope="col">Privilegio</th>
                                <th scope="col">Grupo</th>
                                <th scope="col">Biometría</th>
                            </tr>
                        </thead>
                        <tbody>
                            {% for employee in employees %}
                            <tr class="{% if employee.uid in selected %}selected-row{% endif %}">
                                <td>
                                    <div class="form-check">
                                        <input class="form-check-input" type="checkbox" name="selected" value="{{ employee.uid }}" {% if employee.uid in selected %}checked{% endif %}>
                                    </div>
                                </td>
                                <td>{{ employee.uid }}</td>
                                <td>{{ employee.name }}</td>
                                <td>{{ employee.user_id }}</td>
                                <td>{{ employee.card }}</td>
                                <td>{{ employee.privilege }}</td>
                                <td>{{ employee.group_id }}</td>
                                <td>
                                    {% if employee.biometrics %}
                                    <ul class="biometric-list">
                                        {% for bio in employee.biometrics %}
                                        <li>FID {{ bio.fid }}, Tipo {{ bio.type }}, Tamaño {{ bio.size or 'N/D' }}, Válido {{ bio.valid }}</li>
                                        {% endfor %}
                                    </ul>
                                    {% else %}
                                    <span class="text-muted">Sin datos</span>
                                    {% endif %}
                                </td>
                            </tr>
                            {% endfor %}
                        </tbody>
                    </table>
                </form>
            </div>
        </div>
        {% elif terminal %}
            <div class="alert alert-info" role="alert">No se encontraron empleados para el terminal {{ terminal }}.</div>
        {% endif %}
    </main>
    <script src="https://code.jquery.com/jquery-3.7.1.min.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/js/bootstrap.bundle.min.js"></script>
    <script src="https://cdn.datatables.net/1.13.6/js/jquery.dataTables.min.js"></script>
    <script src="https://cdn.datatables.net/1.13.6/js/dataTables.bootstrap5.min.js"></script>
    <script>
        document.addEventListener('DOMContentLoaded', function () {
            const selectionInfo = document.getElementById('selection-info');
            const tableElement = document.getElementById('employees-table');

            if (!tableElement) {
                if (selectionInfo) {
                    selectionInfo.dataset.total = '0';
                    selectionInfo.dataset.selected = '0';
                    selectionInfo.textContent = 'Seleccionados: 0 de 0';
                }
                return;
            }

            const $table = $('#employees-table');
            const dataTable = $table.DataTable({
                pageLength: 50,
                order: [[1, 'asc']],
                autoWidth: false,
                columnDefs: [
                    { orderable: false, targets: 0 },
                ],
                language: {
                    decimal: ',',
                    thousands: '.',
                    emptyTable: 'Sin datos disponibles en la tabla',
                    info: 'Mostrando _START_ a _END_ de _TOTAL_ empleados',
                    infoEmpty: 'Mostrando 0 a 0 de 0 empleados',
                    infoFiltered: '(filtrado de _MAX_ registros totales)',
                    lengthMenu: 'Mostrar _MENU_ registros',
                    loadingRecords: 'Cargando...',
                    processing: 'Procesando...',
                    search: 'Buscar:',
                    zeroRecords: 'No se encontraron coincidencias',
                    paginate: {
                        first: 'Primero',
                        last: 'Último',
                        next: 'Siguiente',
                        previous: 'Anterior'
                    },
                },
                dom: '<"row mb-3"<"col-sm-12 col-md-6"l><"col-sm-12 col-md-6 text-md-end"f>>tip',
            });

            const getAllCheckboxes = () => $table.find('tbody input[type="checkbox"]');

            const updateSelectionInfo = () => {
                if (!selectionInfo) {
                    return;
                }
                const selectedCount = getAllCheckboxes().filter(':checked').length;
                const totalEmployees = selectionInfo.dataset.total ? parseInt(selectionInfo.dataset.total, 10) : getAllCheckboxes().length;
                selectionInfo.dataset.selected = String(selectedCount);
                selectionInfo.textContent = `Seleccionados: ${selectedCount} de ${totalEmployees}`;
            };

            const updateRowClasses = () => {
                getAllCheckboxes().each(function () {
                    const $row = $(this).closest('tr');
                    $row.toggleClass('selected-row', this.checked);
                });
            };

            let lastClicked = null;

            $table.on('change', 'tbody input[type="checkbox"]', function () {
                updateRowClasses();
                updateSelectionInfo();
            });

            $table.on('click', 'tbody input[type="checkbox"]', function (event) {
                const checkboxList = getAllCheckboxes().toArray();
                if (event.shiftKey && lastClicked && lastClicked !== this) {
                    const currentIndex = checkboxList.indexOf(this);
                    const lastIndex = checkboxList.indexOf(lastClicked);
                    if (currentIndex !== -1 && lastIndex !== -1) {
                        const start = Math.min(currentIndex, lastIndex);
                        const end = Math.max(currentIndex, lastIndex);
                        const shouldCheck = this.checked;
                        for (let i = start; i <= end; i += 1) {
                            checkboxList[i].checked = shouldCheck;
                        }
                        updateRowClasses();
                        updateSelectionInfo();
                    }
                }
                lastClicked = this;
            });

            $table.on('draw.dt', function () {
                updateRowClasses();
            });

            updateRowClasses();
            updateSelectionInfo();
        });
    </script>
</body>
</html>
"""


@app.route("/", methods=["GET", "POST"])
def index():
    terminal_value = request.values.get("terminal")
    parsed_ip, parsed_port = parse_terminal_value(terminal_value)
    fallback_ip = (request.values.get("ip", "") or "").strip() or None
    ip = parsed_ip or fallback_ip
    port = parsed_port if parsed_ip else _coerce_port(request.values.get("port"))

    employees = []
    selected = set()

    def redirect_with_terminal():
        terminal_param = format_terminal_value(ip, port)
        if terminal_param:
            return redirect(url_for("index", terminal=terminal_param))
        fallback_terminal = (terminal_value or "").strip()
        if fallback_terminal:
            return redirect(url_for("index", terminal=fallback_terminal))
        return redirect(url_for("index"))

    if request.method == "POST":
        action = request.form.get("action")
        if action == "import" and ip:
            try:
                imported_employees = parse_employee_file(request.files.get("employee_file"))
            except ValueError as exc:
                flash(str(exc))
            else:
                _TERMINAL_EMPLOYEES[ip] = imported_employees
                _SELECTED_EMPLOYEES[ip] = set()
                flash(
                    f"Se importaron {len(imported_employees)} empleado(s) en memoria para el terminal {ip}."
                )
            return redirect_with_terminal()
        elif action == "import":
            flash("Debes indicar un terminal para importar empleados.")
            return redirect_with_terminal()
        elif action == "fetch" and ip:
            try:
                employees = fetch_employees(ip, port)
            except Exception as exc:  # pragma: no cover - dependiente del dispositivo
                logger.exception("Error al consultar empleados del terminal %s", ip)
                flash(f"No se pudo obtener la información del terminal {ip}: {exc}")
            else:
                _TERMINAL_EMPLOYEES[ip] = employees
                selected = _SELECTED_EMPLOYEES.setdefault(ip, set())
            return redirect_with_terminal()
        elif action == "select" and ip:
            selected_uids = set(request.form.getlist("selected"))
            _SELECTED_EMPLOYEES[ip] = selected_uids
            return redirect_with_terminal()
        elif action == "delete" and ip:
            selected_uids = set(request.form.getlist("selected"))
            if not selected_uids:
                flash("Selecciona al menos un empleado para eliminar.")
                return redirect_with_terminal()

            cached_employees = _TERMINAL_EMPLOYEES.get(ip, [])
            to_delete = [emp for emp in cached_employees if emp.get("uid") in selected_uids]

            if not to_delete:
                flash(
                    "Los empleados seleccionados no están disponibles en caché. Consulta nuevamente el terminal."
                )
                return redirect_with_terminal()

            try:
                deleted, errors = delete_employees(ip, to_delete, port=port)
            except Exception as exc:  # pragma: no cover - dependiente del dispositivo
                logger.exception("Error al eliminar empleados del terminal %s", ip)
                flash(f"No se pudieron eliminar los empleados seleccionados: {exc}")
            else:
                if deleted:
                    flash(f"Se eliminaron {len(deleted)} empleado(s) del terminal.")
                    _TERMINAL_EMPLOYEES[ip] = [
                        emp for emp in cached_employees if emp.get("uid") not in deleted
                    ]
                    selected_cache = _SELECTED_EMPLOYEES.get(ip, set())
                    selected_cache.difference_update(deleted)
                    _SELECTED_EMPLOYEES[ip] = selected_cache
                if errors:
                    for uid, message in errors:
                        flash(f"No se pudo eliminar el empleado {uid}: {message}")
            return redirect_with_terminal()
        elif action in {"export_csv", "export_json", "export_excel"} and ip:
            selected_uids = set(request.form.getlist("selected"))
            if not selected_uids:
                flash("Selecciona al menos un empleado para exportar.")
                return redirect_with_terminal()

            cached_employees = _TERMINAL_EMPLOYEES.get(ip, [])
            if not cached_employees:
                flash(
                    "No hay empleados en caché para exportar. Consulta primero el terminal."
                )
                return redirect_with_terminal()

            selected_employees = [
                emp for emp in cached_employees if emp.get("uid") in selected_uids
            ]
            if not selected_employees:
                flash(
                    "Los empleados seleccionados no están disponibles en caché. Consulta nuevamente el terminal."
                )
                return redirect_with_terminal()

            _SELECTED_EMPLOYEES[ip] = selected_uids
            export_format = action.split("_", 1)[1]
            try:
                return build_export_response(ip, selected_employees, export_format)
            except ValueError as exc:
                flash(str(exc))
                return redirect_with_terminal()
        elif action == "clear":
            if ip:
                cached = _TERMINAL_EMPLOYEES.pop(ip, [])
                removed_count = len(cached)
                _SELECTED_EMPLOYEES.pop(ip, None)
                if removed_count:
                    flash(
                        f"Se limpiaron {removed_count} empleado(s) en memoria para el terminal {ip}."
                    )
                else:
                    flash("No había empleados almacenados en memoria para este terminal.")
                return redirect_with_terminal()
            _TERMINAL_EMPLOYEES.clear()
            _SELECTED_EMPLOYEES.clear()
            flash("Se limpiaron los empleados almacenados en memoria.")
            return redirect(url_for("index"))

    if ip and ip in _TERMINAL_EMPLOYEES:
        employees = _TERMINAL_EMPLOYEES[ip]
        selected = _SELECTED_EMPLOYEES.get(ip, set())
    
    employee_map = {emp["uid"]: emp for emp in employees}
    employee_uids = set(employee_map.keys())
    selected = {uid for uid in selected if uid in employee_uids}
    if ip:
        _SELECTED_EMPLOYEES[ip] = selected
    total_employees = len(employees)
    selected_count = len(selected)
    terminal_display = format_terminal_value(ip, port)
    if not terminal_display and terminal_value:
        terminal_display = terminal_value.strip()
    return render_template_string(
        INDEX_TEMPLATE,
        ip=ip,
        port=port,
        terminal=terminal_display,
        employees=employees,
        selected=selected,
        employee_map=employee_map,
        total_employees=total_employees,
        selected_count=selected_count,
    )


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
